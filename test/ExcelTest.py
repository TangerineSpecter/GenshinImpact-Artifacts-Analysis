from openpyxl import Workbook

# # 假设有一个字段映射，将字段名与表头对应起来
# field_mapping = {
#     'name': 'Name1',
#     'city': 'City1',
#     'age': 'Age1'
# }
#
# # 假设有一个数据集
# data = [
#     {'name': 'Alice', 'age': 25, 'city': 'London'},
#     {'name': 'Bob', 'age': 30, 'city': 'Paris'},
#     {'name': 'Charlie', 'age': 35, 'city': 'Berlin'}
# ]
#
# # 创建一个新的Excel工作簿
# wb = Workbook()
# ws = wb.active
#
# # 添加表头
# for col, header in enumerate(field_mapping.values(), start=1):
#     cell = ws.cell(row=1, column=col)
#     cell.value = header
#
# # 根据字段映射将数据插入到相应的单元格中
# for row, row_data in enumerate(data, start=2):
#     for field, col in field_mapping.items():
#         cell = ws.cell(row=row, column=list(field_mapping.keys()).index(field) + 1)
#         cell.value = row_data[field]
#
# # 保存Excel文件
# wb.save("data.xlsx")

import pandas as pd
import json

df = pd.read_excel("data.xlsx")
# print(df)

json_data = []
for index, row in df.iterrows():
    role = {
        'name': row['角色'],
        'health': row['大生命'],
        'attack': row['大攻击'],
        'defense': row['大防御'],
        'critical_rate': row['暴击'],
        'critical_damage': row['暴击伤害'],
        'elemental_mastery': row['元素精通'],
        'energy_recharge': row['充能效率']
    }
    # print(row)
    json_data.append(role)
print(json.dumps(json_data, ensure_ascii=False))
